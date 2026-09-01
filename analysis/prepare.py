"""Ekstrak kolom yang dipakai dari xlsx sumber menjadi CSV ringkas."""
import openpyxl, csv, warnings
warnings.filterwarnings("ignore")

KEEP = ["KODE_GARDU","NAMA_GARDU","UP3","ULP_KODE","GARDU_INDUK","TRAFO_GI_KE","PENYULANG","KAPASITAS_KVA",
"IRATA_SIANG_A","IRATA_MALAM_A","IMAX_SIANG_A","IMAX_MALAM_A","I_NOMINAL_A",
"PERSEN_BEBAN_SIANG","PERSEN_BEBAN_MALAM","PERSEN_BEBAN","PERSEN_BEBAN_PUNCAK_FASA","KATEGORI_BEBAN",
"UNBALANCE_MAX_PCT","STATUS_UNBALANCE","LATITUDE","LONGITUDE","STATUS_KOORDINAT","TYPE_GARDU","KOTA_KAB",
"GI_KODE","GI_METODE_MATCH","GI_DAYA_MVA","GI_PERSEN_SIANG","GI_PERSEN_MALAM","GI_UP3","FLAG_ANOMALI"]

wb = openpyxl.load_workbook("Gardu_Beban_Lokasi_JABAR.xlsx", read_only=True)
ws = wb["DATA_GARDU"]; it = ws.iter_rows(values_only=True)
hdr = list(next(it)); idx = [hdr.index(k) for k in KEEP]
n = 0
with open("analysis/gardu.csv", "w", newline="") as f:
    w = csv.writer(f); w.writerow(KEEP)
    for r in it:
        if r[0] is None: continue
        w.writerow(["" if r[i] is None else r[i] for i in idx]); n += 1
wb.close(); print("gardu rows:", n)

wb = openpyxl.load_workbook("Master SPKLU Maret 2026.xlsx", read_only=True)
ws = wb.worksheets[0]; it = ws.iter_rows(values_only=True); hdr = list(next(it))
n = 0
with open("analysis/spklu.csv", "w", newline="") as f:
    w = csv.writer(f); w.writerow(hdr)
    for r in it:
        if r[2] is None: continue
        w.writerow(["" if v is None else v for v in r]); n += 1
wb.close(); print("spklu rows:", n)
