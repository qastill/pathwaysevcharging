import pandas as pd, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "Detail Transaksi SPKLU 202603.xlsx"
OUT = "Rekap Konsumsi Energi per SPKLU Jawa Barat - Maret 2026.xlsx"

# ---- Klasifikasi jenis titik lokasi (identik dgn dashboard index.html) ----
def classify(n):
    s = (n or "").upper()
    def rx(p): return re.search(p, s) is not None
    if rx(r"REST AREA|RA KM|KM \d+ ?[AB]\b|RUAS|TRAVOY| TOL |TOL$|CIPULARANG|CIPALI|PALIKANCI|JAGORAWI|KANCI|PEJAGAN|PADALARANG - CILEUNYI|JAKARTA - CIKAMPEK|CIKAMPEK - PADALARANG|ALVACHARGE|AMANAH RAHARJO"): return "Rest Area Tol"
    if rx(r"MALL|SUMMARECON|LIPPO|PLAZA|Q SQUARE|LIFESTYLE|TRANS STUDIO|PASAR SEGAR|PVJ|PARIS VAN JAVA|CIHAMPELAS|FESTIVAL CITYLINK|\bTSM\b|\bBIP\b|CITYLINK|TRANSMART|TECHNOMART|GALUH MAS|SUPERMARKET|HYPERMART|CARREFOUR|SUPERINDO|GIANT|HERO|LOTTE|RAMAYANA|MATAHARI"): return "Mall / Retail / Supermarket"
    if rx(r"HOTEL|ASTON|\bINN\b|RESORT|DORMITORY|GUEST|HOSTEL|HORISON|HARRIS"): return "Hotel / Hospitality"
    if rx(r"BYD|ARISTA|CHERY|GEELY|WULING|HYUNDAI|TOYOTA|DEALER|SHOWROOM|MARKETING GALLERY|SALES CENTER|AVANTE|\bAUTO\b|MITSUBISHI|NETA|DFSK|\bMG \b"): return "Dealer / Showroom Mobil"
    if rx(r"COFFEE|CAFE|CAFÉ|CAR WASH|RESTO|RESTAURANT|\bKOPI\b|EATERY|FOOD PARK|\bFOOD\b|TERAS|DRUMS|AVENUE|POPI|\bRM \b|WAROENG|WARUNG|CIBIUK|SAMOLO"): return "F&B / Kafe / Leisure Retail"
    if rx(r"\bRS |RSUD|RUMAH SAKIT|MEDIKA|HOSPITAL|KLINIK|SILOAM|HERMINA|SENTOSA|ADVENT|AYSHA"): return "Rumah Sakit / Kesehatan"
    if rx(r"BALAI KOTA|SEKDA|PEMDA|PEMKOT|PEMKAB|DPRD|DISHUB|DINAS|SAMSAT|KECAMATAN|GEDUNG SATE|POLRES|POLSEK|POLDA|POSPOL|KODIM|KEJAKSAAN|PENGADILAN|BANDARA|AIRPORT|STASIUN|STATION|TERMINAL|KAMPUS|UNIVERSITAS|\bITB\b|UNPAD|SEKOLAH|ALUN-ALUN|ALUN ALUN|MASJID|MESJID|ISTANA"): return "Publik / Pemerintah / Transportasi"
    if rx(r"PLN UP3|PLN ULP|PLN UID|PLN UIP|PLN UPT|CSE PLN|\bUP3\b|\bULP\b|\bUPT\b|UPDL|KANWIL|DAYA\+ PLN|DAYA\+ CENTER|DAYA\+ ULP|ICON ?HUB|ICON ?PLUS|\bPOSKO\b|\bUID\b|PLN CGE|MOBILE UID|GARDU INDUK|BISNIS CENTER|PLN 3 |SENTRA KIIC|ASTAKAIA"): return "Kantor / Unit Layanan PLN"
    if rx(r"SPBU|PERTAMINA|SHELL| BP |\bVIVO\b"): return "SPBU / Fuel Station"
    if rx(r"GEOWISATA|WISATA|PUNCAK|PANTAI|BEACH|CURUG|FLOATING|\bDAM\b|JATILUHUR|GLAMPING|VILLA|SAFARI|GOLF|CAMPSITE|CAMP\b|FOREST|RENGGANIS|OCEAN VIEW|PARADISO|FARM"): return "Wisata / Leisure"
    if rx(r"KOTA BARU PARAHYANGAN|\bKBP\b|GRAND DEPOK CITY|GARDENS|BALE PARE|TOWNSHIP|RESIDENCE|\bPERUM\b|CLUSTER|CITRALAND|CITRAGRAN|CITRA |VERONA|HILLS|GRAND TARUMA|MAHATA|FAMILY PARK|PARAHYANGAN|APARTEMEN|APARTMENT|HARVEST CITY|MILLENNIUM CITY|GREENLAND|SAKURA|PANCAR GARDEN|CIANJUR ASRI|GATEWAY|\bKP \b|KAMPUNG"): return "Perumahan / Township"
    if rx(r"BTN|BANK|BSPACE|OFFICE|GEDUNG|TOWER|GRHA|GRAHA|KANTOR|TENTH AVENUE|WALKING"): return "Perkantoran / Gedung Komersial"
    return "Lainnya / Campuran"

def tagging(trx):
    if trx > 50: return "Hijau (Utilisasi Tinggi >50 trx)"
    if trx >= 11: return "Kuning (Utilisasi Sedang 11-50 trx)"
    return "Merah (Utilisasi Rendah <10 trx)"

# ---- Baca & agregasi ----
df = pd.read_excel(SRC, sheet_name="Detail Transaksi")
df = df[df['Provinsi'].astype(str).str.upper() == "JAWA BARAT"].copy()
df['kwh'] = pd.to_numeric(df['kwh'], errors='coerce').fillna(0)
df['rppakai'] = pd.to_numeric(df['rppakai'], errors='coerce').fillna(0)

def daya_num(x):
    m = re.search(r"[\d.]+", str(x));  return float(m.group()) if m else 0

agg = df.groupby(['IdSpklu','Nama SPKLU'], dropna=False).agg(
    Jumlah_Transaksi=('Id Transaksi','count'),
    Energi_Konsumsi_kWh=('kwh','sum'),
    Total_Pendapatan=('rppakai','sum'),
    UP3=('UP3','first'),
    Kota_Kab=('Pemda/Pemkot','first'),
).reset_index()

agg['Jenis Titik Lokasi'] = agg['Nama SPKLU'].apply(classify)
agg['Tagging SPKLU'] = agg['Jumlah_Transaksi'].apply(tagging)
agg['Energi_Konsumsi_kWh'] = agg['Energi_Konsumsi_kWh'].round(2)
agg = agg.sort_values('Energi_Konsumsi_kWh', ascending=False).reset_index(drop=True)
agg.insert(0,'No', range(1,len(agg)+1))

# reorder columns
recap = agg[['No','IdSpklu','Nama SPKLU','Tagging SPKLU','Jenis Titik Lokasi',
             'UP3','Kota_Kab','Jumlah_Transaksi','Energi_Konsumsi_kWh','Total_Pendapatan']].copy()
recap.columns = ['No','ID SPKLU','Nama SPKLU','Tagging SPKLU','Jenis Titik Lokasi',
                 'UP3','Kota/Kabupaten','Jumlah Transaksi','Energi Konsumsi Total (kWh)','Total Pendapatan (Rp)']

# ---- Sheet ringkasan per jenis lokasi ----
by_loc = agg.groupby('Jenis Titik Lokasi').agg(
    Jumlah_SPKLU=('IdSpklu','nunique'),
    Total_kWh=('Energi_Konsumsi_kWh','sum'),
    Total_Transaksi=('Jumlah_Transaksi','sum'),
).reset_index().sort_values('Total_kWh', ascending=False)
by_loc['Rata2 kWh per SPKLU'] = (by_loc['Total_kWh']/by_loc['Jumlah_SPKLU']).round(1)
by_loc['Total_kWh'] = by_loc['Total_kWh'].round(1)
by_loc.columns = ['Jenis Titik Lokasi','Jumlah SPKLU','Total Energi (kWh)','Total Transaksi','Rata-rata kWh per SPKLU']

by_tag = agg.groupby('Tagging SPKLU').agg(
    Jumlah_SPKLU=('IdSpklu','nunique'),
    Total_kWh=('Energi_Konsumsi_kWh','sum'),
    Total_Transaksi=('Jumlah_Transaksi','sum'),
).reset_index()
by_tag['Total_kWh'] = by_tag['Total_kWh'].round(1)
by_tag.columns = ['Tagging SPKLU','Jumlah SPKLU','Total Energi (kWh)','Total Transaksi']

# ---- Tulis ke Excel dgn styling ----
with pd.ExcelWriter(OUT, engine='openpyxl') as xw:
    recap.to_excel(xw, sheet_name='Rekap per SPKLU', index=False, startrow=2)
    by_loc.to_excel(xw, sheet_name='Ringkasan per Jenis Lokasi', index=False, startrow=2)
    by_tag.to_excel(xw, sheet_name='Ringkasan per Tagging', index=False, startrow=2)

# styling
wb = openpyxl.load_workbook(OUT)
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUB_FONT = Font(italic=True, size=10, color="555555")
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin,right=thin,top=thin,bottom=thin)
TAG_FILL = {"Merah":"F8CBAD","Kuning":"FFF2CC","Hijau":"C6E0B4"}

titles = {
 'Rekap per SPKLU': ('Rekap Konsumsi Energi per SPKLU — Jawa Barat (Maret 2026)',
    'Sumber: Detail Transaksi SPKLU periode 202603 · Diurutkan dari energi konsumsi tertinggi'),
 'Ringkasan per Jenis Lokasi': ('Ringkasan Konsumsi Energi per Jenis Titik Lokasi — Jawa Barat (Maret 2026)',
    'Jenis lokasi diklasifikasi dari nama SPKLU'),
 'Ringkasan per Tagging': ('Ringkasan Konsumsi Energi per Tagging Utilisasi — Jawa Barat (Maret 2026)',
    'Tagging: Merah <10 trx · Kuning 11-50 trx · Hijau >50 trx (basis: transaksi Maret 2026)'),
}

for ws in wb.worksheets:
    t, sub = titles[ws.title]
    ws['A1'] = t; ws['A1'].font = TITLE_FONT
    ws['A2'] = sub; ws['A2'].font = SUB_FONT
    hdr_row = 3
    ncol = ws.max_column
    for c in range(1, ncol+1):
        cell = ws.cell(hdr_row, c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    # data borders + number formats + tagging color
    tag_col = None; kwh_cols=[]; rp_cols=[]
    for c in range(1, ncol+1):
        h = ws.cell(hdr_row,c).value
        if h=="Tagging SPKLU": tag_col=c
        if h and "kWh" in str(h): kwh_cols.append(c)
        if h and ("Rp" in str(h) or "Pendapatan" in str(h)): rp_cols.append(c)
        if h in ("Jumlah Transaksi","Total Transaksi","Jumlah SPKLU"): rp_cols  # int
    for r in range(hdr_row+1, ws.max_row+1):
        for c in range(1, ncol+1):
            cell = ws.cell(r,c); cell.border=BORDER
        for c in kwh_cols: ws.cell(r,c).number_format='#,##0.00'
        for c in rp_cols: ws.cell(r,c).number_format='#,##0'
        if tag_col:
            v = str(ws.cell(r,tag_col).value or "")
            for k,color in TAG_FILL.items():
                if v.startswith(k): ws.cell(r,tag_col).fill=PatternFill("solid",fgColor=color)
    # int cols formatting
    for c in range(1,ncol+1):
        if ws.cell(hdr_row,c).value in ("Jumlah Transaksi","Total Transaksi","Jumlah SPKLU"):
            for r in range(hdr_row+1, ws.max_row+1): ws.cell(r,c).number_format='#,##0'
    ws.freeze_panes = ws.cell(hdr_row+1,1)
    # column widths
    widths = {'No':6,'ID SPKLU':10,'Nama SPKLU':46,'Tagging SPKLU':30,'Jenis Titik Lokasi':30,
              'UP3':16,'Kota/Kabupaten':20,'Jumlah Transaksi':14,'Energi Konsumsi Total (kWh)':18,
              'Total Pendapatan (Rp)':18,'Jenis Titik Lokasi ':30,'Jumlah SPKLU':13,
              'Total Energi (kWh)':16,'Total Transaksi':14,'Rata-rata kWh per SPKLU':18}
    for c in range(1,ncol+1):
        h = ws.cell(hdr_row,c).value
        ws.column_dimensions[get_column_letter(c)].width = widths.get(h,16)
    ws.row_dimensions[hdr_row].height=30

wb.save(OUT)

# ringkasan cetak
tot_kwh = recap['Energi Konsumsi Total (kWh)'].sum()
print(f"SPKLU unik      : {len(recap):,}")
print(f"Total transaksi : {recap['Jumlah Transaksi'].sum():,}")
print(f"Total energi    : {tot_kwh:,.1f} kWh")
print(f"File            : {OUT}")
print("\nTop 5 SPKLU by kWh:")
print(recap.head(5)[['ID SPKLU','Nama SPKLU','Tagging SPKLU','Jenis Titik Lokasi','Energi Konsumsi Total (kWh)']].to_string(index=False))
print("\nRingkasan per jenis lokasi:")
print(by_loc.to_string(index=False))
