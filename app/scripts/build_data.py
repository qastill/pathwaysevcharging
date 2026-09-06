#!/usr/bin/env python3
"""Build the app's data payloads from the raw workbooks in the repository root.

Outputs (app/public/data/):
  stations.json  - every public charging station in Indonesia (PLN + non-PLN), enriched
                   for West Java with brand, connector mix, price, utilisation and hourly pattern
  hosts.json     - peer-to-peer home chargers, derived from the *distribution* of West Java
                   home-charging installations; coordinates jittered, identities pseudonymised
  meta.json      - tariffs, hourly patterns by venue type, national monthly totals, assumptions

Run from the app/ directory:  python3 scripts/build_data.py
"""
import csv, json, math, os, random, re, sys, collections
import openpyxl
import pandas as pd

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
OUT = os.path.join(os.path.dirname(__file__), "..", "public", "data")
os.makedirs(OUT, exist_ok=True)
P = lambda *a: os.path.join(ROOT, *a)

random.seed(2026)

ENERGY_RP = 2466.75            # PLN SPKLU energy tariff, Rp/kWh (from 101k March-2026 transactions)
DEFAULT_PPJ = 0.08             # street-lighting levy where the pemda is unknown
NONPLN_EST_RP = 3200           # indicative all-in price at non-PLN operators (flagged as estimate)

def norm(s): return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())
def kw_of(s):
    m = re.search(r"[\d.,]+", str(s or "")); 
    return float(m.group().replace(",", ".")) if m else 0.0
def title(s):
    s = str(s or "").strip()
    return s if s.isupper() is False else s.title()

# ---------------------------------------------------------------- venue classifier (shared with dashboard)
def classify(n):
    s = (n or "").upper()
    rx = lambda p: re.search(p, s) is not None
    if rx(r"REST AREA|RA KM|KM \d+ ?[AB]\b|RUAS|TRAVOY| TOL |TOL$|CIPULARANG|CIPALI|PALIKANCI|JAGORAWI|KANCI|PEJAGAN|PADALARANG - CILEUNYI|JAKARTA - CIKAMPEK|CIKAMPEK - PADALARANG|ALVACHARGE|AMANAH RAHARJO"): return "rest_area"
    if rx(r"MALL|SUMMARECON|LIPPO|PLAZA|Q SQUARE|LIFESTYLE|TRANS STUDIO|PASAR SEGAR|PVJ|PARIS VAN JAVA|CIHAMPELAS|FESTIVAL CITYLINK|\bTSM\b|\bBIP\b|CITYLINK|TRANSMART|TECHNOMART|GALUH MAS|SUPERMARKET|HYPERMART|CARREFOUR|SUPERINDO|GIANT|HERO|LOTTE|RAMAYANA|MATAHARI|AEON|GRAND INDONESIA|SENAYAN|PONDOK INDAH|CENTRAL PARK|KOTA KASABLANKA|PIK|PANTAI INDAH"): return "mall"
    if rx(r"HOTEL|ASTON|\bINN\b|RESORT|DORMITORY|GUEST|HOSTEL|HORISON|HARRIS|SANTIKA|IBIS|NOVOTEL|MERCURE|SWISS|HILTON|MARRIOTT|SHERATON|HYATT|WESTIN"): return "hotel"
    if rx(r"BYD|ARISTA|CHERY|GEELY|WULING|HYUNDAI|TOYOTA|DEALER|SHOWROOM|MARKETING GALLERY|SALES CENTER|AVANTE|\bAUTO\b|MITSUBISHI|NETA|DFSK|\bMG \b|AION|JAECOO|VINFAST|HONDA|NISSAN|LEXUS|BMW|MERCEDES|DENZA|XPENG|ZEEKR|POLYTRON"): return "dealer"
    if rx(r"COFFEE|CAFE|CAFÉ|CAR WASH|RESTO|RESTAURANT|\bKOPI\b|EATERY|FOOD PARK|\bFOOD\b|TERAS|DRUMS|AVENUE|POPI|\bRM \b|WAROENG|WARUNG|CIBIUK|SAMOLO|STARBUCKS|MCD|KFC"): return "fnb"
    if rx(r"\bRS |RSUD|RUMAH SAKIT|MEDIKA|HOSPITAL|KLINIK|SILOAM|HERMINA|SENTOSA|ADVENT|AYSHA"): return "hospital"
    if rx(r"BALAI KOTA|SEKDA|PEMDA|PEMKOT|PEMKAB|DPRD|DISHUB|DINAS|SAMSAT|KECAMATAN|GEDUNG SATE|POLRES|POLSEK|POLDA|POSPOL|KODIM|KEJAKSAAN|PENGADILAN|BANDARA|AIRPORT|STASIUN|STATION|TERMINAL|KAMPUS|UNIVERSITAS|\bITB\b|UNPAD|SEKOLAH|ALUN-ALUN|ALUN ALUN|MASJID|MESJID|ISTANA|KEMENTERIAN|KANTOR GUBERNUR|WALIKOTA|BUPATI|PELABUHAN|PARKIR"): return "public"
    if rx(r"PLN UP3|PLN ULP|PLN UID|PLN UIP|PLN UPT|CSE PLN|\bUP3\b|\bULP\b|\bUPT\b|UPDL|KANWIL|DAYA\+ PLN|DAYA\+ CENTER|DAYA\+ ULP|ICON ?HUB|ICON ?PLUS|\bPOSKO\b|\bUID\b|PLN CGE|MOBILE UID|GARDU INDUK|BISNIS CENTER|PLN 3 |SENTRA KIIC|ASTAKAIA|PLN "): return "pln"
    if rx(r"SPBU|PERTAMINA|SHELL| BP |\bVIVO\b"): return "fuel"
    if rx(r"GEOWISATA|WISATA|PUNCAK|PANTAI|BEACH|CURUG|FLOATING|\bDAM\b|JATILUHUR|GLAMPING|VILLA|SAFARI|GOLF|CAMPSITE|CAMP\b|FOREST|RENGGANIS|OCEAN VIEW|PARADISO|FARM|TAMAN|ANCOL|CANDI"): return "leisure"
    if rx(r"KOTA BARU PARAHYANGAN|\bKBP\b|GRAND DEPOK CITY|GARDENS|BALE PARE|TOWNSHIP|RESIDENCE|\bPERUM\b|CLUSTER|CITRALAND|CITRAGRAN|CITRA |VERONA|HILLS|GRAND TARUMA|MAHATA|FAMILY PARK|PARAHYANGAN|APARTEMEN|APARTMENT|HARVEST CITY|MILLENNIUM CITY|GREENLAND|SAKURA|PANCAR GARDEN|CIANJUR ASRI|GATEWAY|\bKP \b|KAMPUNG|BSD|ALAM SUTERA|GADING SERPONG|LIPPO CIKARANG|JABABEKA|PODOMORO|ESTATE"): return "residential"
    if rx(r"BTN|BANK|BSPACE|OFFICE|GEDUNG|TOWER|GRHA|GRAHA|KANTOR|TENTH AVENUE|WALKING|SCBD|WISMA|MENARA|PLAZA"): return "office"
    return "other"

# ---------------------------------------------------------------- coarse province lookup for non-Jabar rows
PROV = [("Aceh",5.55,95.32),("Sumatera Utara",3.59,98.67),("Sumatera Barat",-0.95,100.35),("Riau",0.51,101.45),("Kepulauan Riau",1.08,104.03),
("Jambi",-1.61,103.61),("Sumatera Selatan",-2.99,104.76),("Bangka Belitung",-2.13,106.11),("Bengkulu",-3.80,102.26),("Lampung",-5.43,105.26),
("Banten",-6.12,106.15),("DKI Jakarta",-6.21,106.85),("Jawa Barat",-6.91,107.61),("Jawa Tengah",-6.97,110.42),("DI Yogyakarta",-7.80,110.36),
("Jawa Timur",-7.25,112.75),("Bali",-8.65,115.22),("Nusa Tenggara Barat",-8.58,116.12),("Nusa Tenggara Timur",-10.18,123.60),
("Kalimantan Barat",-0.03,109.34),("Kalimantan Tengah",-2.21,113.92),("Kalimantan Selatan",-3.32,114.59),("Kalimantan Timur",-0.50,117.15),("Kalimantan Utara",2.84,117.37),
("Sulawesi Utara",1.49,124.84),("Gorontalo",0.54,123.06),("Sulawesi Tengah",-0.90,119.87),("Sulawesi Barat",-2.67,118.89),("Sulawesi Selatan",-5.15,119.43),("Sulawesi Tenggara",-3.97,122.51),
("Maluku",-3.70,128.18),("Maluku Utara",0.79,127.38),("Papua Barat",-0.86,134.06),("Papua",-2.53,140.72)]
# sub-regional anchors (city -> province) to sharpen borders around Jabodetabek / Java
ANCH = [("Jakarta","DKI Jakarta",-6.21,106.85),("Tangerang","Banten",-6.18,106.63),("Tangerang Selatan","Banten",-6.29,106.71),("Serang","Banten",-6.12,106.15),("Cilegon","Banten",-6.00,106.05),
("Bekasi","Jawa Barat",-6.24,107.00),("Depok","Jawa Barat",-6.40,106.82),("Bogor","Jawa Barat",-6.60,106.80),("Bandung","Jawa Barat",-6.91,107.61),("Cirebon","Jawa Barat",-6.73,108.55),
("Karawang","Jawa Barat",-6.31,107.30),("Sukabumi","Jawa Barat",-6.92,106.93),("Tasikmalaya","Jawa Barat",-7.33,108.22),("Semarang","Jawa Tengah",-6.97,110.42),("Solo","Jawa Tengah",-7.57,110.83),
("Tegal","Jawa Tengah",-6.87,109.14),("Yogyakarta","DI Yogyakarta",-7.80,110.36),("Surabaya","Jawa Timur",-7.25,112.75),("Malang","Jawa Timur",-7.98,112.63),("Denpasar","Bali",-8.65,115.22),
("Medan","Sumatera Utara",3.59,98.67),("Palembang","Sumatera Selatan",-2.99,104.76),("Makassar","Sulawesi Selatan",-5.15,119.43),("Balikpapan","Kalimantan Timur",-1.27,116.83),("Batam","Kepulauan Riau",1.08,104.03),
("Pekanbaru","Riau",0.51,101.45),("Lampung","Lampung",-5.43,105.26),("Padang","Sumatera Barat",-0.95,100.35),("Banjarmasin","Kalimantan Selatan",-3.32,114.59),("Manado","Sulawesi Utara",1.49,124.84),
("Pontianak","Kalimantan Barat",-0.03,109.34),("Mataram","Nusa Tenggara Barat",-8.58,116.12),("Samarinda","Kalimantan Timur",-0.50,117.15),("Jayapura","Papua",-2.53,140.72),("Kupang","Nusa Tenggara Timur",-10.18,123.60)]
def region_of(lat, lng):
    best = None
    for name, prov, la, lo in ANCH:
        d = (lat-la)**2 + (lng-lo)**2
        if best is None or d < best[0]: best = (d, name, prov)
    if best[0] < 0.35**2: return best[1], best[2]
    bp = min(PROV, key=lambda p: (lat-p[1])**2 + (lng-p[2])**2)
    return None, bp[0]

# ---------------------------------------------------------------- 1. national master
wb = openpyxl.load_workbook(P("SPKLU_Indonesia_Lengkap_2026-06-08.xlsx"), read_only=True)
rows = list(wb["Master SPKLU Indonesia"].iter_rows(values_only=True))
H = rows[0]; idx = {h: i for i, h in enumerate(H)}
stations = {}
for r in rows[1:]:
    sid = str(r[idx["ID SPKLU"]]).strip()
    try: lat, lng = float(r[idx["Latitude"]]), float(r[idx["Longitude"]])
    except: continue
    if not (-11 < lat < 6 and 95 < lng < 141): continue
    st = {
        "id": sid, "name": str(r[idx["Nama SPKLU"]] or "").strip(), "address": str(r[idx["Alamat"]] or "").strip(),
        "lat": round(lat, 6), "lng": round(lng, 6),
        "kw": kw_of(r[idx["Kapasitas (kW)"]]), "status": str(r[idx["Status"]] or "").strip().lower(),
        "chargers": int(r[idx["Jumlah Charger"]] or 0), "connectors": int(r[idx["Jumlah Connector"]] or 0),
        "operator": "PLN" if r[idx["Kategori"]] == "PLN" else "Mitra",
    }
    stations[sid] = st
print("national stations:", len(stations))

# national monthly totals
monthly = []
for sname in ["Konsumsi kWh 2024-2026"]:
    for r in list(wb[sname].iter_rows(values_only=True))[1:]:
        if r[0] and r[1]: monthly.append({"year": int(r[0]), "month": str(r[1]).title(), "trx": int(r[2] or 0), "kwh": round(float(r[3] or 0)), "rp": round(float(r[4] or 0))})

# ---------------------------------------------------------------- 2. West Java master (brand, AC/DC, connectors, address, city)
wb2 = openpyxl.load_workbook(P("Master SPKLU Maret 2026.xlsx"), read_only=True)
jr = list(wb2.active.iter_rows(values_only=True)); JH = {h: i for i, h in enumerate(jr[0])}
by_name = {norm(s["name"]): s for s in stations.values()}
jabar_units = collections.defaultdict(list)
for r in jr[1:]:
    st = by_name.get(norm(r[JH["LOKASI"]]))
    if not st: continue
    jabar_units[st["id"]].append(r)
for sid, units in jabar_units.items():
    st = stations[sid]
    u0 = units[0]
    st["address"] = str(u0[JH["ALAMAT"]] or st["address"]).strip() or st["address"]
    st["city"] = str(u0[JH["KOTA/KAB"]] or "").strip().title().replace("Kab. ", "Kab. ").replace("Kota ", "Kota ")
    st["province"] = "Jawa Barat"
    st["up3"] = str(u0[JH["UP3"]] or "").title()
    st["operator"] = "PLN" if any(u[JH["MILIK"]] == "PLN" for u in units) else "Mitra"
    brands = sorted({re.sub(r"^\d+\.\s*", "", str(u[JH["MEREK"]] or "")).strip() for u in units if u[JH["MEREK"]]})
    st["brands"] = brands
    st["units"] = [{"brand": re.sub(r"^\d+\.\s*", "", str(u[JH["MEREK"]] or "")).strip(), "kw": kw_of(u[JH["KW"]]), "type": u[JH["OUTPUT"]] or ("DC" if kw_of(u[JH["KW"]]) >= 25 else "AC"), "n": int(u[JH["JML KONEKTOR"]] or 1), "since": (u[JH["TGL OPERASI"]].year if hasattr(u[JH["TGL OPERASI"]], "year") else None)} for u in units]
    st["kw"] = max(st["kw"], max(x["kw"] for x in st["units"]))
    yrs = [x["since"] for x in st["units"] if x["since"]]
    if yrs: st["since"] = min(yrs)
print("jabar-enriched:", len(jabar_units))

# ---------------------------------------------------------------- 3. Data SPKLU.xlsx plug list (real plug types for ~100 PLN Jabar units)
wb3 = openpyxl.load_workbook(P("Data SPKLU.xlsx"), read_only=True)
plug_by_coord = []
for r in list(wb3["Sheet1"].iter_rows(values_only=True))[2:]:
    if r[10] and r[13]:
        m = re.findall(r"-?\d+\.\d+", str(r[13]))
        if len(m) >= 2: plug_by_coord.append((float(m[0]), float(m[1]), str(r[10])))

# ---------------------------------------------------------------- 4. recap (utilisation March 2026) + transactions (hourly, duration, connectors, price)
rec = list(csv.DictReader(open(P("Rekap_SPKLU_Jabar_ArcGIS.csv"), encoding="utf-8-sig")))
for r in rec:
    sid = str(r["ID_SPKLU"]).zfill(5); st = stations.get(sid)
    if not st: continue
    st["city"] = st.get("city") or str(r["Kota_Kab"]).title()
    st["province"] = "Jawa Barat"
    st["venueLabel"] = r["Jenis_Lokasi"]
    st["stats"] = {"trx": int(r["Jml_Transaksi"]), "kwh": round(float(r["Energi_kWh"])), "rp": int(float(r["Pendapatan_Rp"])),
                   "tag": "high" if "Hijau" in r["Tagging_SPKLU"] else ("mid" if "Kuning" in r["Tagging_SPKLU"] else "low")}

df = pd.read_csv(P("Detail Transaksi SPKLU Jawa Barat - Maret 2026.csv"), low_memory=False)
df["sid"] = df["ID SPKLU"].astype(str).str.zfill(5)
df["kwh"] = pd.to_numeric(df["Energi (kWh)"], errors="coerce").fillna(0)
df["min"] = pd.to_numeric(df["Durasi (menit)"], errors="coerce").fillna(0)
df["rp"] = pd.to_numeric(df["Total Bayar (Rp)"], errors="coerce").fillna(0)
df["dow"] = pd.to_datetime(df["Tanggal"]).dt.dayofweek
df = df[(df.kwh > 0.5) & (df["min"] > 2) & (df["min"] < 600)]

def conn_type(s):
    s = str(s).upper()
    if "CHADEMO" in s: return "CHAdeMO"
    if "CCS" in s: return "CCS2"
    if "TYPE 2" in s or "TYPE2" in s or "AC" in s: return "Type 2"
    return "CCS2"

venue_hours = collections.defaultdict(lambda: [0]*24)
venue_meta = {}
for sid, g in df.groupby("sid"):
    st = stations.get(sid)
    if not st: continue
    hours = [0]*24
    for h, c in g["Jam"].value_counts().items(): hours[int(h)] = int(c)
    dow = [0]*7
    for d, c in g["dow"].value_counts().items(): dow[int(d)] = int(c)
    conns = collections.defaultdict(lambda: {"n": 0, "kw": 0.0, "trx": 0})
    for (cname, dkw), cg in g.groupby(["Connector", "Daya Charger"]):
        t = conn_type(cname); k = kw_of(re.search(r"(\d+)\s*kW", str(cname)).group(1) if re.search(r"(\d+)\s*kW", str(cname)) else dkw)
        key = (t, k); conns[key]["n"] += 1; conns[key]["kw"] = k; conns[key]["trx"] += len(cg)
    st["plugs"] = [{"type": t, "kw": k, "n": v["n"], "trx": v["trx"]} for (t, k), v in sorted(conns.items(), key=lambda x: -x[1]["kw"])]
    st["usage"] = {"n": int(len(g)), "hours": hours, "dow": dow,
                   "avgMin": round(float(g["min"].median()), 1), "avgKwh": round(float(g["kwh"].median()), 1),
                   "p90Min": round(float(g["min"].quantile(0.9)), 1)}
    tot = g["rp"].sum(); kw = g["kwh"].sum()
    if kw > 0: st["priceKwh"] = int(round(tot / kw / 10.0) * 10)
    v = classify(st["name"]); vh = venue_hours[v]
    for i in range(24): vh[i] += hours[i]

# ---------------------------------------------------------------- 5. finalise every station
ppj = json.load(open(P("analysis", "price.json")))["pemda"]
ppj_rate = {p["pemda"].upper(): p["rate"] / 100.0 for p in ppj}
def plugs_from_kw(st):
    n = max(1, st["connectors"] or st["chargers"] or 1)
    if st["kw"] >= 25:
        return [{"type": "CCS2", "kw": st["kw"], "n": n, "trx": 0}]
    return [{"type": "Type 2", "kw": st["kw"] or 22, "n": n, "trx": 0}]

out = []
for st in stations.values():
    if "province" not in st:
        city, prov = region_of(st["lat"], st["lng"]); st["province"] = prov
        if city: st["city"] = city
    st["venue"] = classify(st["name"])
    st["type"] = "DC" if st["kw"] >= 25 else "AC"
    if "plugs" not in st:
        st["plugs"] = plugs_from_kw(st)
        # real plug lists for the surveyed PLN units
        for la, lo, plugs in plug_by_coord:
            if abs(la - st["lat"]) < 0.0006 and abs(lo - st["lng"]) < 0.0006:
                types = [("CHAdeMO" if "CHADEMO" in p else "CCS2" if "CCS2" in p else "Type 2") for p in re.split(r"[;&]", plugs.upper())]
                st["plugs"] = [{"type": t, "kw": (st["kw"] if t != "Type 2" else min(st["kw"], 22)), "n": 1, "trx": 0} for t in types]
                break
    if "priceKwh" not in st:
        if st["operator"] == "PLN":
            rate = ppj_rate.get(str(st.get("city", "")).upper(), DEFAULT_PPJ)
            st["priceKwh"] = int(round(ENERGY_RP * (1 + rate) / 10.0) * 10)
        else:
            st["priceKwh"] = NONPLN_EST_RP; st["priceEstimated"] = True
    st["speed"] = "ultra" if st["kw"] >= 100 else "fast" if st["kw"] >= 50 else "medium" if st["kw"] >= 20 else "slow"
    # reliability score: utilisation + status + age, 0..5
    base = 3.6
    if st.get("stats"): base += {"high": 1.0, "mid": 0.5, "low": 0.0}[st["stats"]["tag"]]
    if st["status"] in ("available", "inuse"): base += 0.3
    if st["status"] in ("maintenance", "unavailable"): base -= 1.2
    if st["status"] == "offline mode": base -= 0.4
    st["score"] = round(max(1.0, min(5.0, base)), 1)
    st["reviews"] = int(min(500, (st.get("stats", {}).get("trx", 0) / 10)))
    st["amenities"] = {"rest_area": ["toilet", "musala", "food", "parking"], "mall": ["food", "toilet", "wifi", "parking"], "hotel": ["food", "wifi", "parking", "toilet"],
                       "pln": ["parking", "toilet", "musala"], "fnb": ["food", "wifi", "toilet"], "dealer": ["parking", "wifi", "toilet"], "office": ["parking", "wifi"],
                       "public": ["parking", "toilet"], "hospital": ["parking", "toilet", "food"], "fuel": ["toilet", "musala", "food"], "leisure": ["parking", "food"],
                       "residential": ["parking"], "other": ["parking"]}[st["venue"]]
    st["open"] = "24h" if st["venue"] in ("rest_area", "pln", "fuel", "hospital", "residential", "public", "other") else "08:00–22:00"
    out.append(st)

for s in out: s.pop("units", None)
out.sort(key=lambda s: (-(s.get("stats", {}).get("trx", 0)), s["id"]))
json.dump(out, open(os.path.join(OUT, "stations.json"), "w"), ensure_ascii=False, separators=(",", ":"))
print("stations written:", len(out), "with usage:", sum(1 for s in out if "usage" in s), "with stats:", sum(1 for s in out if "stats" in s))

# ---------------------------------------------------------------- 6. P2P hosts from the spatial distribution of home-charging installs (pseudonymised)
ev = pd.read_csv(P("Data pelanggan EV.txt"), sep="\t", low_memory=False, encoding="latin-1")
ev = ev.dropna(subset=["Lat", "Long"])
ev = ev[(ev.Lat < -5) & (ev.Lat > -8) & (ev.Long > 105) & (ev.Long < 109.5)]
ev = ev[ev["Status Approval"] == "Selesai"]
FIRST = ["Adi","Agus","Andi","Ayu","Bagus","Bayu","Budi","Citra","Dedi","Dewi","Dian","Dimas","Dwi","Eka","Fajar","Fitri","Galih","Gita","Hadi","Hendra","Ika","Indra","Intan","Irfan","Joko","Kartika","Lestari","Lina","Maya","Nanda","Nur","Putri","Rahmat","Ratna","Reza","Rina","Rizky","Sari","Siti","Taufik","Tri","Wahyu","Widya","Yoga","Yuli","Yusuf","Zainal","Asep","Cecep","Dadang","Euis","Neneng","Ujang","Yayat"]
def kelurahan_name(s):
    s = str(s or ""); return s.split(" - ", 1)[1].title() if " - " in s else s.title()
hosts = []
# stratified sample: at most 3 hosts per kelurahan to keep the map balanced, ~360 total
sample = ev.sample(frac=1, random_state=7).groupby("Kelurahan").head(2)
sample = sample.sample(min(len(sample), 380), random_state=11)
for i, (_, r) in enumerate(sample.iterrows()):
    daya = float(r["tarif/Daya"] or 7700)
    kw = 7.0 if daya < 11000 else 11.0 if daya < 22000 else 22.0
    # jitter 250–600 m so a listing never points at a real address
    ang = random.random() * 2 * math.pi; dist = random.uniform(0.25, 0.6)
    lat = float(r["Lat"]) + (dist / 111.0) * math.sin(ang); lng = float(r["Long"]) + (dist / (111.0 * math.cos(math.radians(-6.5)))) * math.cos(ang)
    kel = kelurahan_name(r["Kelurahan"]); kec = kelurahan_name(r["Kecamatan"]); kab = kelurahan_name(r["Kabupaten"])
    fname = random.choice(FIRST)
    price = random.choice([2200, 2300, 2400, 2500, 2500, 2600, 2700, 2800])
    windows = random.choice([["18:00", "06:00"], ["08:00", "17:00"], ["00:00", "24:00"], ["20:00", "07:00"], ["09:00", "21:00"]])
    days = random.choice(["daily", "daily", "weekdays", "weekends"])
    brand = str(r["Jenis Kendaraan"] or "").strip()
    hosts.append({
        "id": f"h{i+1:04d}", "name": f"Charger {fname}", "host": f"{fname} {random.choice('ABCDHKMNPRSTW')}.",
        "lat": round(lat, 5), "lng": round(lng, 5), "area": (kel if kel == kec else f"{kel}, {kec}"), "city": kab, "province": "Jawa Barat",
        "kw": kw, "plug": "Type 2", "priceKwh": price, "days": days, "from": windows[0], "to": windows[1],
        "score": round(random.uniform(4.3, 5.0), 1), "reviews": random.randint(2, 48), "sessions": random.randint(5, 120),
        "vehicle": brand, "amenities": random.sample(["covered", "cctv", "wifi", "toilet", "coffee", "musala", "wide"], random.randint(1, 3)),
        "note": random.choice(["Carport tertutup, bisa ditinggal.", "Colokan di garasi samping rumah, akses gerbang otomatis.", "Tersedia kursi tunggu dan air minum.", "Lokasi dekat gerbang cluster, mudah dicari.", "Charger wallbox terpasang resmi, sudah SLO.", "Bisa ngecas semalaman, ambil pagi."]),
        "since": random.choice([2024, 2024, 2025, 2025, 2026]), "simulated": True,
    })
json.dump(hosts, open(os.path.join(OUT, "hosts.json"), "w"), ensure_ascii=False, separators=(",", ":"))
print("hosts written:", len(hosts))

# ---------------------------------------------------------------- 7. meta
price = json.load(open(P("analysis", "price.json")))
hours_all = [0]*24
for v, hs in venue_hours.items():
    for i in range(24): hours_all[i] += hs[i]
meta = {
    "builtFrom": "PLN UID Jawa Barat detail transaksi Maret 2026 (101.020 trx), Master SPKLU Indonesia 8 Juni 2026, Master SPKLU Jabar Maret 2026",
    "tariff": {"energyRp": ENERGY_RP, "ppjDefault": DEFAULT_PPJ, "avgAllIn": round(price["price"]["allin"]), "ppjByCity": {p["pemda"].title(): round(p["rate"], 1) for p in ppj}},
    "hoursByVenue": {v: hs for v, hs in venue_hours.items()}, "hoursAll": hours_all,
    "session": {"medianMin": round(float(df["min"].median()), 1), "medianKwh": round(float(df["kwh"].median()), 1), "meanKwh": round(float(df["kwh"].mean()), 1)},
    "assume": price["assume"], "carbon": price["carbon"], "cost": price["cost"], "monthly": monthly,
    "counts": {"stations": len(out), "pln": sum(1 for s in out if s["operator"] == "PLN"), "mitra": sum(1 for s in out if s["operator"] != "PLN"),
               "dc": sum(1 for s in out if s["type"] == "DC"), "provinces": len({s["province"] for s in out}), "hosts": len(hosts)},
    "ev_brands": {k: int(v) for k, v in ev["Jenis Kendaraan"].value_counts().head(12).items()},
}
json.dump(meta, open(os.path.join(OUT, "meta.json"), "w"), ensure_ascii=False, separators=(",", ":"))
print("meta written", meta["counts"])
